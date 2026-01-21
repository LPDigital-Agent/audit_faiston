"""
File Reader Tools for FileAnalyzer Agent

Provides tools for reading CSV, XLSX, and XLS files from S3.
Following the LLM=Brain / Python=Hands principle from CLAUDE.md:
- Python handles I/O, parsing, and data extraction
- LLM analyzes the extracted content and makes mapping decisions
"""

import csv
import io
import json
import logging
import unicodedata
from typing import Dict, List, Any, Optional

import boto3

from shared.debug_utils import debug_error

logger = logging.getLogger(__name__)

# S3 bucket for inventory documents
DEFAULT_BUCKET = "faiston-one-sga-documents-prod"


def read_file_from_s3(
    s3_key: str,
    bucket: str = DEFAULT_BUCKET,
    max_rows: int = 100,
) -> Dict[str, Any]:
    """
    Read file content from S3 for analysis.

    Supports CSV, XLSX, and XLS files. Returns structured data
    for LLM analysis including headers, sample rows, and metadata.

    Args:
        s3_key: S3 object key (path to file)
        bucket: S3 bucket name
        max_rows: Maximum number of rows to read for analysis

    Returns:
        Dict containing:
        - success: bool
        - file_type: str (csv, xlsx, xls)
        - headers: List[str]
        - rows: List[List[str]]
        - row_count: int
        - column_count: int
        - error: Optional[str]
    """
    try:
        s3 = boto3.client("s3")
        # NFC normalize S3 key to match upload encoding
        # Prevents NoSuchKey errors with Portuguese characters (Ç, Ã, Õ)
        normalized_key = unicodedata.normalize("NFC", s3_key)
        logger.info("[FileReader] Reading file from s3://%s/%s", bucket, normalized_key)

        response = s3.get_object(Bucket=bucket, Key=normalized_key)
        content = response["Body"].read()

        # Detect file type from extension
        file_type = _detect_file_type(s3_key)

        if file_type == "csv":
            return parse_csv_content(content, max_rows)
        elif file_type in ("xlsx", "xls"):
            return parse_excel_content(content, file_type, max_rows)
        else:
            return {
                "success": False,
                "error": f"Unsupported file type: {file_type}",
                "file_type": file_type,
                "headers": [],
                "rows": [],
                "row_count": 0,
                "column_count": 0,
            }

    except Exception as e:
        debug_error(e, "file_reader_read_s3", {"s3_key": s3_key, "bucket": bucket})
        return {
            "success": False,
            "error": str(e),
            "file_type": "unknown",
            "headers": [],
            "rows": [],
            "row_count": 0,
            "column_count": 0,
        }


def parse_csv_content(
    content: bytes,
    max_rows: int = 100,
    encoding: str = "utf-8",
) -> Dict[str, Any]:
    """
    Parse CSV content and extract headers + sample rows.

    PERF-001: Optimized to avoid loading entire file into memory.
    Uses itertools.islice to read only needed rows, then counts remaining lines.

    Args:
        content: Raw CSV bytes
        max_rows: Maximum rows to return
        encoding: Text encoding (default UTF-8)

    Returns:
        Dict with parsed data
    """
    import itertools

    try:
        # Try UTF-8 first, then Latin-1 as fallback
        try:
            text = content.decode(encoding)
        except UnicodeDecodeError:
            text = content.decode("latin-1")
            logger.info("[FileReader] Fell back to latin-1 encoding")

        # PERF-001: Use itertools.islice to read only first N+1 rows (header + max_rows)
        # This avoids loading 10,000+ rows into memory when we only need 100
        reader = csv.reader(io.StringIO(text))

        # Read header + max_rows sample rows
        first_rows = list(itertools.islice(reader, max_rows + 1))

        if not first_rows:
            return {
                "success": True,
                "file_type": "csv",
                "headers": [],
                "rows": [],
                "row_count": 0,
                "column_count": 0,
            }

        headers = first_rows[0]
        data_rows = first_rows[1:] if len(first_rows) > 1 else []

        # Count remaining rows efficiently (without storing them)
        # sum(1 for _ in reader) counts remaining rows after islice consumed first N+1
        remaining_rows = sum(1 for _ in reader)
        total_rows = len(data_rows) + remaining_rows

        # Convert rows to strings for consistent handling
        data_rows = [[str(cell) for cell in row] for row in data_rows]

        logger.info(
            "[FileReader] Parsed CSV: %d columns, %d total rows, returning %d sample rows",
            len(headers),
            total_rows,
            len(data_rows),
        )

        return {
            "success": True,
            "file_type": "csv",
            "headers": headers,
            "rows": data_rows,
            "row_count": total_rows,
            "column_count": len(headers),
        }

    except Exception as e:
        debug_error(e, "file_reader_parse_csv", {"encoding": encoding})
        return {
            "success": False,
            "error": f"CSV parse error: {e}",
            "file_type": "csv",
            "headers": [],
            "rows": [],
            "row_count": 0,
            "column_count": 0,
        }


def parse_excel_content(
    content: bytes,
    file_type: str,
    max_rows: int = 100,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Parse Excel (XLSX/XLS) content and extract headers + sample rows.

    Args:
        content: Raw Excel bytes
        file_type: 'xlsx' or 'xls'
        max_rows: Maximum rows to return
        sheet_name: Specific sheet to read (default: first sheet)

    Returns:
        Dict with parsed data including sheet info
    """
    try:
        import pandas as pd

        # Read Excel file
        if file_type == "xlsx":
            xlsx = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
        else:  # xls
            xlsx = pd.ExcelFile(io.BytesIO(content), engine="xlrd")

        sheet_names = xlsx.sheet_names

        # Select sheet
        target_sheet = sheet_name or sheet_names[0]
        if target_sheet not in sheet_names:
            return {
                "success": False,
                "error": f"Sheet '{target_sheet}' not found. Available: {sheet_names}",
                "file_type": file_type,
                "headers": [],
                "rows": [],
                "row_count": 0,
                "column_count": 0,
                "sheets": sheet_names,
            }

        # PERF-001: Read the sheet ONCE, get total rows, then truncate
        # Previously called pd.read_excel() twice (once for data, once for count)
        df_full = pd.read_excel(xlsx, sheet_name=target_sheet)
        total_rows = len(df_full)

        # Truncate to max_rows for analysis (keep memory efficient)
        df = df_full.head(max_rows)
        del df_full  # Free memory immediately

        headers = df.columns.tolist()
        # Convert to strings, handle NaN
        rows = df.fillna("").astype(str).values.tolist()

        logger.info(
            "[FileReader] Parsed %s: sheet='%s', %d columns, %d total rows",
            file_type.upper(),
            target_sheet,
            len(headers),
            total_rows,
        )

        return {
            "success": True,
            "file_type": file_type,
            "headers": headers,
            "rows": rows,
            "row_count": total_rows,
            "column_count": len(headers),
            "sheets": sheet_names,
            "active_sheet": target_sheet,
        }

    except ImportError as e:
        debug_error(e, "file_reader_parse_excel_import", {"file_type": file_type})
        return {
            "success": False,
            "error": f"Excel support not available: {e}",
            "file_type": file_type,
            "headers": [],
            "rows": [],
            "row_count": 0,
            "column_count": 0,
        }
    except Exception as e:
        debug_error(e, "file_reader_parse_excel", {"file_type": file_type, "sheet_name": sheet_name})
        return {
            "success": False,
            "error": f"Excel parse error: {e}",
            "file_type": file_type,
            "headers": [],
            "rows": [],
            "row_count": 0,
            "column_count": 0,
        }


def _detect_file_type(s3_key: str) -> str:
    """Detect file type from S3 key extension."""
    lower_key = s3_key.lower()
    if lower_key.endswith(".csv"):
        return "csv"
    elif lower_key.endswith(".xlsx"):
        return "xlsx"
    elif lower_key.endswith(".xls"):
        return "xls"
    elif lower_key.endswith(".pdf"):
        return "pdf"
    elif lower_key.endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp")):
        return "image"
    return "unknown"


def get_column_statistics(rows: List[List[str]], headers: List[str]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate basic statistics for each column.

    Args:
        rows: Data rows
        headers: Column headers

    Returns:
        Dict mapping column name to statistics
    """
    stats = {}

    for col_idx, header in enumerate(headers):
        values = [row[col_idx] if col_idx < len(row) else "" for row in rows]

        non_empty = [v for v in values if v.strip()]
        unique_values = set(non_empty)

        stats[header] = {
            "total_count": len(values),
            "null_count": len(values) - len(non_empty),
            "unique_count": len(unique_values),
            "sample_values": list(unique_values)[:5],
            "inferred_type": _infer_column_type(non_empty),
        }

    return stats


def _infer_column_type(values: List[str]) -> str:
    """Infer column data type from sample values."""
    if not values:
        return "unknown"

    # Check if all values are numeric
    numeric_count = 0
    date_count = 0

    for v in values[:20]:  # Sample first 20
        try:
            float(v.replace(",", "."))
            numeric_count += 1
        except ValueError:
            pass

        # Simple date pattern check
        if "/" in v or "-" in v:
            parts = v.replace("-", "/").split("/")
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                date_count += 1

    sample_size = min(len(values), 20)
    if numeric_count > sample_size * 0.8:
        return "number"
    if date_count > sample_size * 0.8:
        return "date"
    return "string"


# =============================================================================
# Data Extraction Functions (Moved from gemini_text_analyzer.py)
# =============================================================================
# These functions are pure Python - no LLM calls.
# They extract full data rows using column mappings for import execution.


async def extract_data_from_file(
    s3_key: str,
    column_mappings: Dict[str, str],
    bucket: str = DEFAULT_BUCKET,
    max_rows: int = 5000,
) -> Dict[str, Any]:
    """
    Extract and transform data from file using column mappings.

    This function reads the file from S3 and extracts rows
    transformed according to the column mappings.
    Pure Python implementation - no LLM calls.

    Args:
        s3_key: S3 key where file is stored
        column_mappings: Validated mappings {source_column: target_field}
        bucket: S3 bucket name
        max_rows: Maximum rows to extract

    Returns:
        {
            "success": bool,
            "rows": [...],
            "row_count": int,
            "errors": [...],
        }
    """
    logger.info("[FileReader] Extracting data from: %s", s3_key)

    try:
        import unicodedata

        s3 = boto3.client("s3")
        # NFC normalize S3 key to match how files were uploaded
        # Prevents NoSuchKey errors with Portuguese characters (Ç, Ã, Õ)
        normalized_key = unicodedata.normalize("NFC", s3_key)
        response = s3.get_object(Bucket=bucket, Key=normalized_key)
        content = response["Body"].read()

        file_type = _detect_file_type(s3_key)

        rows = []
        errors = []

        if file_type == "csv":
            rows, errors = _extract_csv_rows(content, column_mappings, max_rows)
        elif file_type == "xlsx":
            rows, errors = _extract_xlsx_rows(content, column_mappings, max_rows)
        elif file_type == "xls":
            rows, errors = _extract_xls_rows(content, column_mappings, max_rows)
        else:
            return {
                "success": False,
                "error": f"Formato nao suportado: {file_type}",
                "rows": [],
            }

        logger.info("[FileReader] Extracted %d rows, %d errors", len(rows), len(errors))

        return {
            "success": True,
            "rows": rows,
            "row_count": len(rows),
            "errors": errors[:50],  # Limit errors
            "errors_count": len(errors),
        }

    except Exception as e:
        debug_error(e, "file_reader_extract_data", {"s3_key": s3_key, "bucket": bucket})
        return {
            "success": False,
            "error": str(e),
            "rows": [],
        }


def _extract_csv_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform CSV rows using mappings."""
    # Detect encoding
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    # Detect delimiter
    sample = text[:4096]
    delimiters = [",", ";", "\t", "|"]
    delimiter_counts = {d: sample.count(d) for d in delimiters}
    delimiter = max(delimiter_counts, key=delimiter_counts.get)

    lines = text.strip().split("\n")
    reader = csv.DictReader(lines, delimiter=delimiter)

    rows = []
    errors = []

    for i, row in enumerate(reader):
        if i >= max_rows:
            break

        try:
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row:
                    transformed[target_field] = row[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})

    return rows, errors


def _extract_xlsx_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform XLSX rows using mappings."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)

    if not headers:
        wb.close()
        return [], [{"row": 1, "error": "No headers found"}]

    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

    rows = []
    errors = []

    for i, row_values in enumerate(rows_iter):
        if i >= max_rows:
            break

        try:
            # Build row dict
            row_dict = {}
            for j, val in enumerate(row_values):
                if j < len(headers):
                    row_dict[headers[j]] = val

            # Transform using mappings
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row_dict:
                    transformed[target_field] = row_dict[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})

    wb.close()
    return rows, errors


def _extract_xls_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform XLS rows using mappings."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        return [], [{"row": 1, "error": "Empty sheet"}]

    headers = [str(ws.cell_value(0, c)) or f"Column_{c}" for c in range(ws.ncols)]

    rows = []
    errors = []

    for r in range(1, min(ws.nrows, max_rows + 1)):
        try:
            # Build row dict
            row_dict = {}
            for c in range(ws.ncols):
                row_dict[headers[c]] = ws.cell_value(r, c)

            # Transform using mappings
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row_dict:
                    transformed[target_field] = row_dict[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": r + 1, "error": str(e)})

    return rows, errors
